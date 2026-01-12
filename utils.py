"""
This script contains some convenience functions.
"""
import numpy as np
import pandas as pd
import openpyxl
import sciris as sc
import os
from pgmpy.factors.discrete import TabularCPD
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import styles
from platform import system
from getpass import getuser
from socket import gethostname
from os import path, getlogin
from pgmpy.global_vars import logger
from constants import TODAY_STR, TIME_STR

MEDIUM_BORDER = styles.Side(border_style='medium')
THIN_BORDER = styles.Side(border_style='thin')
DASHDOT_BORDER = styles.Side(border_style='dashDot')
DASHED_BORDER = styles.Side(border_style='dashed')


def get_desktop_folder(subfolder='ACT HR script output'):
    """
    Gets the system specific Desktop folder and appends the subfolder to the path.
    Copied from Optima package.

    :param subfolder: subfolder on the desktop, defaults to 'ACT HR script output', set to None or '' for the Desktop itself
    :return: the path to the subfolder on the Desktop
    """
    import platform
    system = platform.system()
    if subfolder is None:
        subfolder = ''
    if system == 'Windows':
        username = getlogin()
        desktop_folder =  f'C:\\Users\\{username}\\Desktop\\{subfolder}'
        if not desktop_folder.endswith('\\'):
            desktop_folder = desktop_folder + '\\'
    elif system == 'Darwin': #MacOS X
        username = getuser()
        desktop_folder =  f'/Users/{username}/Desktop/{subfolder}'
        if not desktop_folder.endswith('/'):
            desktop_folder = desktop_folder + '/'
    elif system == 'Linux': # Linux we default to the home directory and we use _ in place of spaces on Linux, for ease of typing
        username = getuser()
        desktop_folder =  f'/home/{username}/{subfolder.replace(" ","_")}'
        if not desktop_folder.endswith('/'):
            desktop_folder = desktop_folder + '/'
    else:
        raise Exception(f'Error: location of desktop on {system} not yet implemented.')
    return desktop_folder


def unstack_keep_order(df, level):
    """Unstack one level of index, widening the dataframe."""
    new_index = df.index.droplevel(level).unique()
    new_cols = df.index.get_level_values(level).unique()
    return df.unstack(level).reindex(new_index).reindex(new_cols, level=-1, axis='columns')


# if using this function need to double-check it is doing the right thing
def update_cpd_with_soft_evidence(cpd, evidence, soft_evidence_prob, state_map):
    """
    Update the CPD of a node with soft evidence.

    Parameters:
    - cpd: The original TabularCPD for the node
    - evidence: Dictionary with evidence variables and their values
    - soft_evidence_prob: Probability of the soft evidence

    Returns:
    - Updated TabularCPD
    """
    # Get the original CPD values
    values = cpd.values
    var_card = cpd.variable_card
    evidence_vars = list(evidence.keys())

    # Calculate the index for the evidence variable
    evidence_index = [state_map[var][val] for var, val in evidence.items()]

    # Create a copy of the CPD values
    updated_values = np.copy(values)

    # Update CPD values based on the soft evidence
    for idx in range(len(values)):
        if all((idx % var_card ** (i + 1)) // (var_card ** i) == evidence_index[i] for i in range(len(evidence))):
            updated_values[idx] = soft_evidence_prob
        else:
            # Normalize other values based on the soft evidence
            updated_values[idx] = (1 - soft_evidence_prob) / (var_card - 1)

    # Normalize the updated values
    updated_values /= updated_values.sum()

    return TabularCPD(variable=cpd.variable,
                      variable_card=var_card,
                      values=updated_values.reshape(cpd.variable_card),
                      evidence=cpd.evidence,
                      evidence_card=cpd.evidence_card)


def df_to_excel(df, ws, header=True, index=False, fit_col_width=True, grouping=None, startrow=0, startcol=0):
    """Write DataFrame df to openpyxl worksheet ws"""

    rows = dataframe_to_rows(df, header=header, index=index)

    if grouping is None:
        new_groups = set()
    else:
        if isinstance(grouping, str):
            grouping = [grouping]
        new_groups_bool = df[grouping].ne(df[grouping].shift()).all('columns')
        new_groups = set(new_groups_bool[new_groups_bool].index.values)

    for r_idx, row in enumerate(rows, startrow + 1):
        for c_idx, value in enumerate(row, startcol + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = value
            if r_idx == 1 and header:
                cell.font = styles.Font(bold=True)
                cell.border = styles.Border(top=MEDIUM_BORDER, left=MEDIUM_BORDER, right=MEDIUM_BORDER, bottom=MEDIUM_BORDER)
            elif (r_idx - 1) in new_groups:
                add_border(cell, bottom=THIN_BORDER)

    if fit_col_width:
        excel_fit_col_width(ws)


def excel_styles(ws, style_dict, bottom_border=None):
    cols = ws.iter_cols(max_row=ws.max_row, max_col=ws.max_column)
    for c_idx, col in enumerate(cols, 1):
        col_heading = col[0].value
        col_style = style_dict.get(col_heading)
        if col_style is None:
            if c_idx == ws.max_column:
                col_style = style_dict.get('right')
        if col_style is not None:
            col_border = col_style.get('border')
            col_font = col_style.get('font')
            col_number_format = col_style.get('number_format')
            for r_idx in range(2, ws.max_row+1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if col_border is not None:
                    add_border(cell, **col_border)
                if col_font is not None:
                    cell.font = col_font
                if col_number_format is not None:
                    cell.number_format = col_number_format

    if bottom_border is not None:
        last_row = ws[ws.max_row]
        for cell in last_row:
            add_border(cell, **bottom_border)


def add_border(cell, left=None, right=None, top=None, bottom=None):
    cell.border = styles.Border(left=left if left else cell.border.left,
                                right=right if right else cell.border.right,
                                top=top if top else cell.border.top,
                                bottom=bottom if bottom else cell.border.bottom)


def excel_fit_col_width(ws):
    # Iterate over all columns and adjust their widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 0.92
        ws.column_dimensions[column_letter].width = adjusted_width


def sheets_dict_to_wb(sheets_dict):
    # Save as Excel workbook
    wb = openpyxl.Workbook()

    # Save each sheet
    for sheet_name, sheet in sheets_dict.items():
        ws = wb.create_sheet(sheet_name)
        for r_idx, row in enumerate(sheet, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                if r_idx == 1 or r_idx == 2:
                    cell.font = styles.Font(bold=True)
                    # cell.border = styles.Border(top=MEDIUM_BORDER, left=MEDIUM_BORDER, right=MEDIUM_BORDER, bottom=MEDIUM_BORDER)

        excel_fit_col_width(ws)

        # probs_styles = {'outcome': {'border': {'right': THIN_BORDER}}}
        # excel_styles(ws, probs_styles, bottom_border={'bottom': THIN_BORDER})

    # Write file
    del wb['Sheet']
    return wb


# A helper function to compute probability distributions from simulated samples.
def get_distribution(samples, variables=None):
    """
    For marginal distribution, P(A): get_distribution(samples, variables=['A'])
    For joint distribution, P(A, B): get_distribution(samples, variables=['A', 'B'])
    """
    if variables is None:
        raise ValueError("variables must be specified")

    return samples.groupby(variables).size() / samples.shape[0]


def get_updated_cpt(inference, evidence, levels):
    # Query the full joint distribution of the node and its parents
    logger.disabled = True
    try:
        joint_distribution = inference.query(variables=list(levels.keys()), virtual_evidence=evidence, state_names=levels)
    except:
        joint_distribution = inference.query(variables=list(levels.keys()), virtual_evidence=evidence)
    logger.disabled = False
    joint_distribution_values = joint_distribution.values

    # Reshape the joint distribution into a conditional probability table (CPT)
    with np.errstate(divide='ignore', invalid='ignore'):
        cpt = np.divide(joint_distribution_values, joint_distribution_values.sum(axis=-1, keepdims=True))
    df = array_to_dataframe(cpt, levels)

    return df


def array_to_dataframe(array, levels):
    """Convert multidimensional np.array to pd.DataFrame"""
    index = pd.MultiIndex.from_product(list(levels.values()), names=list(levels.keys()))
    df = pd.DataFrame({'value': array.flatten()}, index=index)['value']
    return df


def update_dataframe(df_existing, df_new, id_cols):
    """Update new DataFrame structure with values from existing DataFrame"""
    non_id_cols = (set(df_existing.columns) | set(df_new.columns)) - set(id_cols)

    df_existing[[col for col in df_new.columns if col not in df_existing.columns]] = None
    df_new[[col for col in df_existing.columns if col not in df_new.columns]] = None

    for col in df_new.columns:
        if df_new[col].dtype != df_existing[col].dtype:
            df_existing[col] = df_existing[col].astype(object)
            df_new[col] = df_new[col].astype(object)

    df_merged = pd.merge(left=df_existing, right=df_new, how='right', on=id_cols, suffixes=('', '_y'))
    df_merged = df_merged.drop(['{}_y'.format(col) for col in non_id_cols], axis='columns')

    return df_merged


def joint2marginal(variables, output_joint):
    """Extract marginal distributions from joint distribution"""
    result_dict = {}
    all_vars = set(variables)
    for var in variables:
        result_dict[var] = output_joint.marginalize(all_vars - {var}, inplace=False).normalize(inplace=False)
    return result_dict


def nest(l, key_indices) -> dict:
    result = {}
    for d in list(l):
        target = result
        # traverse all keys but the last
        for key in key_indices:
            target = target.setdefault(d.pop(key), {})
        target.update(d)
    return result


def df_to_nested_dict(df: pd.DataFrame, indices) -> dict:
    l = df.to_dict(orient='records')
    return nest(l, indices)


_desktop_folder = get_desktop_folder()
def save_safely(save_func, filename, folder=None, append=TIME_STR, doprint=False):
    if isinstance(save_func, pd.DataFrame): save_func = save_func.to_excel
    if isinstance(save_func, openpyxl.Workbook): save_func = save_func.save

    path = sc.makefilepath(filename, folder=folder)
    try:
        save_func(path)
        if doprint: print('Saved: ', path)
        return path
    except PermissionError:
        basename = os.path.basename(path)
        folder = os.path.dirname(path)
        extension = basename.split('.')[-1]
        new_basename = f'{basename[:-len(extension) - 1]}_{append}.{extension}'

        new_path = sc.makefilepath(new_basename, folder=folder)
        save_func(new_path)
        if doprint: print('Saved: ', new_path)
        return new_path


def read_excel_safely(filename, *args, _read_func=None, _create_tmp_copy=True, **kwargs):
    # creates temporary copy of excel file that is currently open
    if _read_func is None: _read_func = pd.read_excel

    try:
        return _read_func(filename, *args, **kwargs)
    except PermissionError as E:
        if not sc.iswindows(): raise E
        if not _create_tmp_copy: raise E

        from tempfile import mkstemp
        import os
        import gc
        import win32file  # pip install pywin32

        handle, temp_filepath = mkstemp(suffix='.' + filename.split('.')[-1])
        win32file.CopyFile(filename, temp_filepath, 0)
        output = _read_func(temp_filepath, *args, **kwargs)
        gc.collect()
        os.close(handle)
        win32file.DeleteFile(temp_filepath)
        return output


def list_vals(series):
    data = series.values
    with np.printoptions(legacy='1.21'):
        str_joined_list = str(list(data))[1:-1]
    return str_joined_list


def list_vals_sorted(series):
    data = sorted(series.values)
    with np.printoptions(legacy='1.21'):
        str_joined_list = str(list(data))[1:-1]
    return str_joined_list


def modified_triangle_distribution(low, best, high, rng):
    """
    Sample from the custom semi-symmetric triangle distribution with median at best.
    """
    size = len(low)
    assert len(low) == len(best) == len(high)
    u = rng.uniform(low=0, high=1, size=size)
    samples = -np.ones_like(u)

    # Parameters
    denom_left = (high - low) * (best - low)
    denom_right = (high - low) * (high - best)

    # For u <= 0.5, sample from left piece
    i_left = u <= 0.5
    samples[i_left] = low[i_left] + np.sqrt(u[i_left] * denom_left[i_left])

    # For u > 0.5, sample from right piece
    i_right = ~i_left
    samples[i_right] = high[i_right] - np.sqrt((1 - u[i_right]) * denom_right[i_right])

    assert np.all(low[i_left] <= samples[i_left] <= best[i_left])
    assert np.all(best[i_right] <= samples[i_right] <= high[i_right])

    return samples


def transformed_beta(low, mean, high, rng, precision=10, ci=None, eps=1e-3):

    if not len(low) == len(mean) == len(high):
        raise AssertionError
    if not precision > 0:
        raise ValueError

    width = high - low

    # # confidence interval method not yet developed
    # if rel_var is None:  # use confidence interval method
    #     if ci is None:
    #         ci = 0.95  # default confidence interval for lower and upper

    mean_norm = (mean - low) / width

    alpha = mean_norm * precision
    beta = (1 - mean_norm) * precision

    if np.any(alpha <= 0) or np.any(beta <= 0):
        print(f"Beta distribution error: alpha = {alpha}, beta = {beta}")
        raise ValueError

    samples_scaled = low + width * rng.beta(a=alpha, b=beta)

    return samples_scaled


def series_quantile(series, quantile):
    return np.quantile(series.values, quantile)


def Q1(series): return series_quantile(series, quantile=0.25)
def Q3(series): return series_quantile(series, quantile=0.75)


def series_outlier_cutoffs(series, m=1.5):
    q1 = Q1(series)
    q3 = Q3(series)
    iqr = q3 - q1
    lb = q1 - m * iqr
    ub = q3 + m * iqr
    return [lb, ub]


def outlier_cutoff_low(series, m=1.5): return series_outlier_cutoffs(series, m=m)[0]
def outlier_cutoff_high(series, m=1.5): return series_outlier_cutoffs(series, m=m)[1]


def list_outliers(series, m=1.5):
    data = series.values
    lb, ub = series_outlier_cutoffs(series, m=m)
    out = data[np.logical_or(data < lb, data > ub)]
    with np.printoptions(legacy='1.21'):
        str_joined_outliers = str(list(out))[1:-1]
    return str_joined_outliers


def min_without_outliers(series, m=1.5):
    data = series.values
    lb, ub = series_outlier_cutoffs(series, m=m)
    return data[np.logical_and(data >= lb, data <= ub)].min()  # within the bounds


def max_without_outliers(series, m=1.5):
    data = series.values
    lb, ub = series_outlier_cutoffs(series, m=m)
    return data[np.logical_and(data >= lb, data <= ub)].max()  # within the bounds


def ci_range(series, k=1.96):
    data = series.values
    lb = np.mean(data) - k * np.std(data)
    ub = np.mean(data) + k * np.std(data)
    return [lb, ub]


def ci95_low(series): return ci_range(series, k=1.96)[0]
def ci95_high(series): return ci_range(series, k=1.96)[1]
def ci95_low_quantile(series): return series_quantile(series, quantile=0.025)
def ci95_high_quantile(series): return series_quantile(series, quantile=0.975)

